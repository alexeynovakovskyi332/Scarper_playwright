import asyncio
import random
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

URL = "https://rozetka.com.ua/ua/mobile-phones/c80003/"  # вставляем нужную ссылку на товары (для Rozetka )
TOTAL_PAGES = 12  # количество страниц которые будут обрабатыватся (менять по нужде)
BROWSERS = 3
PAGES_PER_BROWSER = 3

data = []  # сюда собираем все товары


async def parse_page(context, page_num):
    page = await context.new_page()
    stealth = Stealth()
    await stealth.apply_stealth_async(page)

    url = f"{URL}page={page_num}/"
    print(f"Обрабатываю страницу {page_num}")

    try:
        await page.goto(url)
        await asyncio.sleep(random.uniform(1, 3))

        products = page.locator("rz-product-tile")
        count = await products.count()
        print(f"Найдено товаров: {count}")

        for i in range(count):
            product = products.nth(i)
            try:
                title = await product.locator("a.tile-title").inner_text()
                link = await product.locator("a.tile-title").get_attribute("href")
                price = await product.locator("div.price").inner_text()
                available = "В наличии"
                if await product.locator("button.buy-button").count() == 0:
                    available = "Нет в наличии"

                data.append({
                    "Страница": page_num,
                    "Название": title.strip(),
                    "Цена": price.strip(),
                    "Наличие": available,
                    "Ссылка": link.strip()
                })
            except:
                continue
    except Exception as e:
        print(f"Ошибка страницы {page_num}: {e}")

    await page.close()


async def browser_worker(playwright, pages):
    browser = await playwright.chromium.launch(headless=False)
    context = await browser.new_context()
    tasks = []
    for page_num in pages:
        tasks.append(parse_page(context, page_num))
    await asyncio.gather(*tasks)
    await browser.close()


async def main():
    async with async_playwright() as p:
        pages = list(range(1, TOTAL_PAGES + 1))
        page_groups = [pages[i:i + PAGES_PER_BROWSER] for i in range(0, len(pages), PAGES_PER_BROWSER)]

        for i in range(0, len(page_groups), BROWSERS):
            batch = page_groups[i:i + BROWSERS]
            print(f"Запускаю batch: {batch}")
            tasks = []
            for group in batch:
                tasks.append(browser_worker(p, group))
            await asyncio.gather(*tasks)


if __name__ == "__main__":
    asyncio.run(main())

    # --- сортируем по странице и фильтруем дубликаты по ссылке ---
    sorted_data = sorted(data, key=lambda x: x["Страница"])
    unique_data = {}
    for item in sorted_data:
        unique_data[item["Ссылка"]] = item  # перезаписываем дубликаты

    # --- сохраняем в Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    headers = ["Страница", "Название", "Цена", "Наличие", "Ссылка"]
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = header_font

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_index, item in enumerate(unique_data.values(), start=2):
        ws.cell(row=row_index, column=1, value=item["Страница"])
        ws.cell(row=row_index, column=2, value=item["Название"])
        ws.cell(row=row_index, column=3, value=item["Цена"])
        ws.cell(row=row_index, column=4, value=item["Наличие"])
        ws.cell(row=row_index, column=5, value=item["Ссылка"])

        if item["Наличие"] == "Нет в наличии":
            for col in range(1, 6):
                ws.cell(row=row_index, column=col).fill = red_fill

    # автоширина колонок
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("rozetka_products.xlsx")
    print(f"Собрано товаров: {len(unique_data)}")